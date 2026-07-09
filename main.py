import os
import pandas as pd
from dotenv import load_dotenv
from langchain_openai import AzureChatOpenAI
from langchain.agents import create_agent
from langchain_core.tools import tool
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX

load_dotenv()

if not os.getenv("AZURE_OPENAI_API_KEY"):
    raise ValueError("AZURE_OPENAI_API_KEY not found in .env file")

EXCLUDE_SHEETS = {"Summary"}
DEFAULT_EXCEL_PATH = "HK Master Price List - July 2026 OS_CLEANED.xlsm"
HEADER_ROW = 4

SELECTED_COLUMNS = [
    "Brand",
    "Supplier",
    "Product Category Code",
    "Product Category",
    "Brand Code",
    "Model (Item Code)",
    "Finish",
    "Description",
    "Unit",
    "Project/ Spec Price",
]

PRICE_COLUMN = "Project/ Spec Price"
MODEL_COLUMN = "Model (Item Code)"
FINISH_COLUMN = "Finish"
DESCRIPTION_COLUMN = "Description"

# Hex colors that mean "discontinued" — adjust to match your workbook
DISCONTINUED_COLORS = {
    "FF0000",   # red
    "C00000",   # dark red
    # add more hex values after inspecting your file
}

YELLOW_COLORS = {
    "FFFF00", # common Excel light yellow
    "FFEB9C",   
    "FFC000",
}

# Yellow = field updates only (NOT price)
YELLOW_SCAN_COLUMNS = [
    "Product Category Code",
    "Product Category",
    "Brand Code",      
    MODEL_COLUMN,
    DESCRIPTION_COLUMN,
    FINISH_COLUMN,    
    "Unit"   
    
]

STRING_COLUMNS = [
    "Brand", "Supplier", "Product Category Code", "Brand Code",
    MODEL_COLUMN, "Finish", DESCRIPTION_COLUMN, "Unit", PRICE_COLUMN
]

_cached_df: pd.DataFrame | None = None



# Step 1: Read one sheet

def _resolve_color_hex(color) -> str | None:
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = str(color.rgb).upper()
        if rgb in ("00000000", "FFFFFFFF", "FF000000"):
            return None
        return rgb[-6:]
    if color.type == "indexed" and color.indexed is not None:
        indexed = COLOR_INDEX.get(color.indexed)
        return str(indexed).upper()[-6:] if indexed else None
    if color.type == "theme":
        return f"theme:{color.theme}"
    return None


def _read_sheet_raw(path: str, sheet_name: str) -> pd.DataFrame | None:
    """Read one sheet into a raw DataFrame with values only."""
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=3, engine="openpyxl")
        if df.empty:
            return None
        df = df.copy()
        df.columns = df.columns.astype(str).str.strip()
        df["source_sheet"] = sheet_name
        return df
    except Exception as e:
        print(f"  ❌ {sheet_name}: {e}")
        return None
    
# Step 2: Normalize each sheet to the same schema 

def _normalize_sheet(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Make every sheet use the same columns and basic types."""
    normalized = pd.DataFrame()
    for col in SELECTED_COLUMNS:
        normalized[col] = df[col] if col in df.columns else pd.NA
    normalized["source_sheet"] = sheet_name
    # Strip strings
    for col in STRING_COLUMNS:
        normalized[col] = normalized[col].astype("string").str.strip()
        normalized[col] = normalized[col].replace("", pd.NA)
    # Standardize model code
    normalized[MODEL_COLUMN] = normalized[MODEL_COLUMN].str.upper()
    # Price as numeric
    normalized[PRICE_COLUMN] = pd.to_numeric(normalized[PRICE_COLUMN], errors="coerce")
    # Drop rows with no model and no price
    normalized = normalized.dropna(subset=[MODEL_COLUMN, PRICE_COLUMN], how="all")
    return normalized.reset_index(drop=True)

# ── Step 3: Merge all sheets ────────────────────────────────────────
def _merge_all_sheets(path: str) -> pd.DataFrame | None:
    """Read, normalize, and concatenate all sheets into one table."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_names = [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]
    wb.close()
    print(f"📄 Merging {len(sheet_names)} sheets...")
    frames = []
    for sheet_name in sheet_names:
        raw = _read_sheet_raw(path, sheet_name)
        if raw is None:
            continue
        normalized = _normalize_sheet(raw, sheet_name)
        if normalized.empty:
            print(f"  ⚠️ {sheet_name}: no usable rows")
            continue
        frames.append(normalized)
        print(f"  ✅ {sheet_name}: {len(normalized)} rows")
    if not frames:
        return None
    merged = pd.concat(frames, ignore_index=True)
    print(f"📦 Merged total: {len(merged)} rows from {len(frames)} sheets")
    return merged


# ── Step 4: Enforce consistency on merged data ────────────────────────
def _enforce_consistency(df: pd.DataFrame) -> pd.DataFrame:
    """Apply cross-sheet consistency rules on the merged table."""
    df = df.copy()
    # Re-apply normalization in case concat introduced inconsistencies
    for col in STRING_COLUMNS:
        df[col] = df[col].astype("string").str.strip().str.upper() if col == MODEL_COLUMN \
            else df[col].astype("string").str.strip()
        df[col] = df[col].replace("", pd.NA)
    df[PRICE_COLUMN] = pd.to_numeric(df[PRICE_COLUMN], errors="coerce")
    # Flag duplicate model codes across sheets
    dup_mask = df.duplicated(subset=[MODEL_COLUMN, "Brand", "Finish"], keep=False)
    df["is_duplicate"] = dup_mask
    if dup_mask.any():
        dup_count = dup_mask.sum()
        print(f"  ⚠️ Found {dup_count} duplicate rows (same model/brand/finish)")
    # Keep first occurrence; comment out next line if you want to keep all duplicates
    df = df.drop_duplicates(subset=[MODEL_COLUMN, "Brand", "Finish"], keep="first")
    # Final column order
    return df[
    SELECTED_COLUMNS + ["source_sheet", "is_duplicate"]
    ].reset_index(drop=True)


# Check if a color indicates discontinued status
def _is_discontinued(color_hex: str | None) -> bool:
    if not color_hex or str(color_hex).startswith("theme:"):
        return False
    return str(color_hex).upper()[-6:] in DISCONTINUED_COLORS

# Check if a color indicates changes product information
def _is_yellow(color_hex: str | None) -> bool:
    if not color_hex or str(color_hex).startswith("theme:"):
        return False
    return str(color_hex).upper()[-6:] in YELLOW_COLORS


def _get_updated_fields(color_by_column: dict[str, str | None]) -> str:
    """Return comma-separated column names that have yellow fill."""
    updated = [
        col for col in YELLOW_SCAN_COLUMNS
        if _is_yellow(color_by_column.get(col))
    ]
    return ", ".join(updated) if updated else ""


def _resolve_status(color_by_column: dict[str, str | None]) -> str:
    """Status is only Discontinued (red price) or Active."""
    price_color = color_by_column.get(PRICE_COLUMN)
    if _is_discontinued(price_color):
        return "Discontinued"
    return "Active"


# ── Step 5: Add colors AFTER merge ──────────────────────────────────
def _attach_cell_status(path: str, df: pd.DataFrame) -> pd.DataFrame:
    """Read red on price (discontinued) and yellow on info columns (updated fields)."""
    wb = load_workbook(path, data_only=True, read_only=False)
    color_lookup: dict[tuple, dict[str, str | None]] = {}
    for sheet_name in df["source_sheet"].unique():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = {
            str(cell.value).strip(): idx
            for idx, cell in enumerate(ws[HEADER_ROW], start=1)
            if cell.value is not None
        }
        model_col_idx = headers.get(MODEL_COLUMN)
        if not model_col_idx:
            continue
        # Scan price (red only) + info columns (yellow)
        columns_to_read = {PRICE_COLUMN: headers.get(PRICE_COLUMN)}
        for col in YELLOW_SCAN_COLUMNS:
            if col in headers:
                columns_to_read[col] = headers[col]
        for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
            model = ws.cell(row=row_idx, column=model_col_idx).value
            if model is None:
                continue
            model_key = str(model).strip().upper()
            row_colors: dict[str, str | None] = {}
            for col_name, col_idx in columns_to_read.items():
                if not col_idx:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                fill = cell.fill
                color_hex = None
                if fill and fill.fill_type:
                    color_hex = _resolve_color_hex(fill.fgColor)
                row_colors[col_name] = color_hex
            color_lookup[(sheet_name, model_key)] = row_colors
    wb.close()
    df = df.copy()
    df["Status"] = df.apply(
        lambda r: _resolve_status(
            color_lookup.get((r["source_sheet"], r[MODEL_COLUMN]), {})
        ),
        axis=1,
    )
    df["Updated_Fields"] = df.apply(
        lambda r: _get_updated_fields(
            color_lookup.get((r["source_sheet"], r[MODEL_COLUMN]), {})
        ),
        axis=1,
    )
    return df[
        SELECTED_COLUMNS + ["source_sheet", "Status", "Updated_Fields", "is_duplicate"]
    ].reset_index(drop=True)


# ── Main processing function ────────────────────────────────────────
def process_cleaned_data(path: str) -> pd.DataFrame | None:
    """Merge all sheets, enforce consistency, then identify cell statuses."""
    merged = _merge_all_sheets(path)
    if merged is None:
        print("❌ No sheets were successfully processed")
        return None
    consistent = _enforce_consistency(merged)
    final = _attach_cell_status(path, consistent)
    print(f"✅ Final master table: {len(final)} rows, {len(final.columns)} columns")
    return final

def _get_dataframe(path: str = DEFAULT_EXCEL_PATH) -> pd.DataFrame | None:
    global _cached_df
    if _cached_df is None:
        _cached_df = process_cleaned_data(path)
    return _cached_df

@tool
def load_price_list(path: str = DEFAULT_EXCEL_PATH) -> str:
    """Load, merge, and clean all price list sheets into one consistent table."""
    global _cached_df
    _cached_df = process_cleaned_data(path)
    if _cached_df is None:
        return "No sheets were successfully processed."
    discontinued_count = (_cached_df["Status"] == "Discontinued").sum()
    active_count = (_cached_df["Status"] == "Active").sum()
    with_updates = (_cached_df["Updated_Fields"] != "").sum()
    print("\n--- First 5 rows of merged master table ---")
    print(_cached_df.head(5))
    print("-------------------------------------------\n")
    return (
        f"Merged {len(_cached_df)} rows from {_cached_df['source_sheet'].nunique()} sheets. "
        f"Columns: {', '.join(_cached_df.columns.astype(str))}. "
       
        f"Active: {active_count}, Discontinued: {discontinued_count}, "
        f"Products with updated fields (yellow): {with_updates}"
    )


@tool
def find_model(model_code: str, path: str = DEFAULT_EXCEL_PATH) -> str:
    """Find a product by model/item code in the merged master table."""
    df = _get_dataframe(path)
    if df is None:
        return "Price list not loaded. Call load_price_list first."
    matches = df[df[MODEL_COLUMN] == model_code.strip().upper()]
    if matches.empty:
        return f"No model found for '{model_code}'."
    lines = []
    for _, row in matches.iterrows():
        updated = row.get("Updated_Fields", "")
        updated_text = (
            f"Updated fields: {updated}"
            if updated and str(updated).strip()
            else "Updated fields: none"
        )
        lines.append(
            f"Brand: {row['Brand']} | "
            f"Supplier: {row['Supplier']} | "
            f"Sheet: {row['source_sheet']} | "
            f"Model: {row[MODEL_COLUMN]} | "
            f"Finish: {row[FINISH_COLUMN]} | "
            f"Description: {row[DESCRIPTION_COLUMN]} | "
            f"Price: {row[PRICE_COLUMN]} | "
            f"UOM: {row['Unit']} | "
            f"Status: {row['Status']} | "
            f"{updated_text}"
        )
    return "\n".join(lines)


@tool
def remove_discontinued_models(path: str = DEFAULT_EXCEL_PATH) -> str:
    """Remove all discontinued models from the merged price list and keep only active products."""
    global _cached_df
    df = _get_dataframe(path)
    if df is None:
        return "Price list not loaded. Call load_price_list first."
    if "Status" not in df.columns:
        return "Status column not found. Call load_price_list first to attach discontinued status."
    total_before = len(df)
    discontinued_count = (df["Status"] == "Discontinued").sum()
    if discontinued_count == 0:
        return f"No discontinued models found. All {total_before} rows are Active."
    _cached_df = df[df["Status"] == "Active"].reset_index(drop=True)
    total_after = len(_cached_df)
    print("\n--- First 5 rows after removing discontinued ---")
    print(_cached_df.head(5))
    print("------------------------------------------------\n")
    return (
        f"Removed {discontinued_count} discontinued models. "
        f"Rows before: {total_before}, rows after: {total_after}. "
        f"Remaining: {total_after} active products."
    )

@tool
def list_updated_products(path: str = DEFAULT_EXCEL_PATH) -> str:
    """List products where Model, Description, or other info fields were updated (yellow cells)."""
    df = _get_dataframe(path)
    if df is None:
        return "Price list not loaded. Call load_price_list first."
    updated = df[df["Updated_Fields"] != ""]
    if updated.empty:
        return "No yellow-highlighted field updates found."
    lines = []
    for _, row in updated.head(50).iterrows():
        lines.append(
            f"Model: {row[MODEL_COLUMN]} | "
            f"Updated: {row['Updated_Fields']} | "
            f"Description: {row[DESCRIPTION_COLUMN]} | "
            f"Status: {row['Status']} | "
            f"Sheet: {row['source_sheet']}"
        )
    return "\n".join(lines)

@tool
def export_active_list(
    output_path: str = "active_price_list.xlsx",
    path: str = DEFAULT_EXCEL_PATH,
) -> str:
    """Export the current active (non-discontinued) price list to an Excel file."""
    df = _get_dataframe(path)
    if df is None:
        return "Price list not loaded. Call load_price_list first."
    if "Status" not in df.columns:
        return "Status column missing. Call load_price_list first."
    active_df = df[df["Status"] == "Active"].copy()
    if active_df.empty:
        return "No active products to export."
    active_df.to_excel(output_path, sheet_name="Active", index=False)
    return f"Exported {len(active_df)} active rows to {output_path}"



# ── Agent setup ─────────────────────────────────────────────────────
model = AzureChatOpenAI(
    azure_deployment=os.getenv("DEPLOYMENT_NAME"),
    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    api_version=os.getenv("API_VERSION", "2025-01-01-preview"),
    azure_endpoint=os.getenv(
        "AZURE_OPENAI_ENDPOINT",
        "https://aa-openai-sweden.openai.azure.com/",
    ),
    openai_api_version=os.getenv("OPENAI_API_VERSION", "2025-08-07"),
)
tools = [load_price_list, remove_discontinued_models, list_updated_products, find_model, export_active_list]
agent = create_agent(model, tools=tools)

if __name__ == "__main__":
    response = agent.invoke(
        {
            "messages": [
                (
                    "human",
                    "Load the merged price list, remove all discontinued models, "
                    "then show how many active products remain, "
                    "the count of updated products (yellow cells), "
                    "and export the active list to active_price_list.xlsx.",
                )
            ]
        }
    )
    print("Agent Response:", response["messages"][-1].content)
